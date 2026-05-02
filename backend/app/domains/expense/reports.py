"""Excel report generation for expenses."""

from collections.abc import Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.domains.expense.models import Expense

_CATEGORY_KO: dict[str, str] = {
    "TRANSPORT": "교통",
    "LODGING": "숙박",
    "MEAL": "식사",
    "MINISTRY": "사역",
    "GIFT": "선물",
    "SUPPLIES": "물품",
    "MEDICAL": "의료",
    "MISC": "기타",
}
_STATUS_KO: dict[str, str] = {
    "pending": "검토 대기",
    "approved": "승인",
    "rejected": "반려",
    "reimbursed": "정산완료",
}
_PAYMENT_KO: dict[str, str] = {
    "PERSONAL_CARD": "개인카드",
    "PERSONAL_CASH": "개인현금",
    "CHURCH_CARD": "교회카드",
    "OTHER": "기타",
}

# Ink-black header fill
_HEADER_FILL = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=10)
_BODY_FONT = Font(name="맑은 고딕", size=10)
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")

_HEADERS = [
    ("날짜", 14),
    ("카테고리", 10),
    ("내용", 32),
    ("업체", 18),
    ("금액", 12),
    ("통화", 6),
    ("결제방법", 12),
    ("상태", 10),
    ("비고", 24),
]


def build_expense_xlsx(expenses: Sequence[Expense], team_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "지출 내역"  # type: ignore[assignment]

    # ── 타이틀 행 ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")  # type: ignore[attr-defined]
    title_cell = ws["A1"]
    title_cell.value = f"{team_name} — 지출 내역"
    title_cell.font = Font(bold=True, name="맑은 고딕", size=13)
    title_cell.alignment = _CENTER
    ws.row_dimensions[1].height = 28

    # ── 헤더 행 ───────────────────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    # ── 데이터 행 ─────────────────────────────────────────────────────────
    total_krw = 0

    for row_idx, expense in enumerate(expenses, start=3):
        spent = expense.spent_at.strftime("%Y-%m-%d") if expense.spent_at else ""
        amount_val = float(expense.amount)
        if expense.currency == "KRW":
            total_krw += amount_val

        row_data = [
            spent,
            _CATEGORY_KO.get(expense.category, expense.category),
            expense.description,
            expense.vendor or "",
            amount_val,
            expense.currency,
            _PAYMENT_KO.get(expense.payment_method or "", "") if expense.payment_method else "",
            _STATUS_KO.get(expense.status, expense.status),
            expense.notes or "",
        ]
        alignments = [_CENTER, _CENTER, _LEFT, _LEFT, _RIGHT, _CENTER, _CENTER, _CENTER, _LEFT]
        for col_idx, (val, align) in enumerate(zip(row_data, alignments), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _BODY_FONT
            cell.alignment = align
            cell.border = _BORDER
            if col_idx == 5:  # amount — number format
                cell.number_format = "#,##0.00"

        # Alternate row shade
        if row_idx % 2 == 0:
            shade = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            for col_idx in range(1, len(_HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = shade

    # ── 합계 행 ───────────────────────────────────────────────────────────
    total_row = len(expenses) + 3
    ws.cell(row=total_row, column=1, value="합계").font = Font(
        bold=True, name="맑은 고딕", size=10
    )
    ws.cell(row=total_row, column=1).alignment = _CENTER
    total_cell = ws.cell(row=total_row, column=5, value=total_krw)
    total_cell.font = Font(bold=True, name="맑은 고딕", size=10)
    total_cell.number_format = "#,##0.00"
    total_cell.alignment = _RIGHT
    ws.cell(row=total_row, column=6, value="KRW").alignment = _CENTER

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
